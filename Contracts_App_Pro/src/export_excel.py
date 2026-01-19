from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Tuple


def export_to_excel(data: List[Tuple], headers: List[str], filename: str) -> bool:
    """
    Export data to Excel file with formatting.
    
    Args:
        data: List of tuples containing row data
        headers: List of column headers
        filename: Output filename (should end with .xlsx)
    
    Returns:
        True if successful, False otherwise
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Справка"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        wb.save(filename)
        return True
    
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return False
